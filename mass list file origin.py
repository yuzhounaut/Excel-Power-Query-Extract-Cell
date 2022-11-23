import os 
import openpyxl
import pandas as pd
import numpy as np #added to save count as a number

path=r"D:/PyProject/Outlier/New"
os.chdir(path)

tables = []
for path, dirs, files in os.walk(path):
    for i in files:
        if i.split(".")[1]  == "xlsx":
            tables.append(i)
Name = pd.DataFrame(tables, columns = ['Name'])
print(Name)
    
merged=pd.DataFrame()
for xlsx_files in os.listdir(path):
     if os.path.isfile(xlsx_files):
                myworkbook=openpyxl.load_workbook(xlsx_files)
                worksheet= myworkbook['Sheet1']
                data = pd.DataFrame(data=[worksheet.cell(2, 1).value], columns= ['File'])
                data['Scan'] = np.array([worksheet.cell(4, 1).value])
                data['RT'] = np.array([worksheet.cell(5, 1).value])              
                    # to grab a different cell, change the 4 and 3 
                    # 4 means 4 rows down, 3 is 3 columns over, not indexed at 0
                merged=pd.concat([merged,data], ignore_index=True)
print(merged)

final_merged=pd.merge(Name, merged, how='outer', left_index=True, right_index=True)
final_merged.to_csv("D:/PyProject/Outlier/111.csv", index=False)
print(final_merged)